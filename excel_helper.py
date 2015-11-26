#!/Python27/python
import logging
from openpyxl import Workbook, load_workbook
import pprint

def read_excel(filename, worksheet, target_class):
    wb = load_workbook(filename, use_iterators = True)
    sheet = wb[worksheet]
    entities = []
    header = []

    for row_idx, row in enumerate(sheet.iter_rows()):
        if row_idx == 0:
            header = [r.internal_value.replace(" ","") for r in row]
            continue
        values = {}
        for col_idx, col in enumerate(header):
            values[col] = row[col_idx].internal_value
        entities.append(target_class(**values))
    logging.info("Read %d entries from %s" % (len(entities), filename))
    return entities
    
class DataHolder:
    a = 0
    b = 0
    def __init__(self, **kwargs):
        """
        Assign all properties given by keyword arguments, even those not defined in class
        """
        for arg in kwargs.keys():
            setattr(self, arg, kwargs[arg])
    def __str__(self):
        """
        Print all meaningful (e.g. either strings or numbers) properties of the the class
        """
        return str({a: getattr(self, a) for a in dir(self) if
                    not a.startswith("_") and isinstance(getattr(self, a), basestring) or isinstance(getattr(self, a), float) or isinstance(getattr(self, a), bool)})
#Use
def write_excel(filename, worksheet, entity_list, headers=None):
    if len(entity_list) == 0:
        return

    wb = Workbook(optimized_write=True)
    ws = wb.create_sheet(title=worksheet)
    if headers is None:
        headers = [a for a in dir(entity_list[0]) if not a.startswith("_")]
    ws.append(headers)

    for entity in entity_list:
        row = [getattr(entity, attrib) for attrib in headers]
        ws.append(row)

    wb.save(filename=filename)
                    
write_excel("sheet.xlsx", "Sheet1", [DataHolder(**{"a":"1", "b":"2"}), DataHolder(**{"a":"3", "b":"4"})])
print "Done"